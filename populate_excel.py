import pandas as pd
from fredapi import Fred
from dotenv import load_dotenv
import os
from openpyxl import load_workbook
from openpyxl.styles import numbers
from datetime import datetime, timedelta
import numpy as np

load_dotenv()
fred = Fred(api_key=os.getenv('FRED_API_KEY'))

# Load the workbook
wb = load_workbook(r'C:\Users\daqui\OneDrive\Documents\Targeted Equity Consulting Group\Macro\Macro Data.xlsx')
ws1 = wb['Sheet1']

# Create Sheet2 if it doesn't exist
if 'Sheet2' not in wb.sheetnames:
    ws2 = wb.create_sheet('Sheet2')
    # Add headers for time series data
    ws2['A3'] = 'Government Payrolls'
    ws2['A4'] = 'Private Payrolls'
    ws2['A5'] = 'Unemployment Rate'
    ws2['A6'] = 'Job Openings Rate'
    ws2['A7'] = 'Job Openings Level'
    ws2['A8'] = 'Avg Hours Worked'
    ws2['A13'] = 'Initial Claims 4W MA'
    ws2['A14'] = 'Continuing Claims'
    ws2['A15'] = 'Temp Jobs'
    ws2['A17'] = 'Retail Sales'
    ws2['A18'] = 'Auto Sales'
    ws2['A19'] = 'Real Disposable Income'
    ws2['A20'] = 'Credit Card Delinquency'
    ws2['A21'] = 'Consumer Sentiment'
    ws2['A22'] = 'Housing Inventory'
    ws2['A23'] = 'Case-Shiller Index'
    ws2['A24'] = 'Median Home Price'
    ws2['A26'] = 'ISM Mfg Employment'
    ws2['A27'] = 'Mfg New Orders'
    ws2['A28'] = 'Industrial Production Mfg'
    ws2['A32'] = 'Household Deposits'
    ws2['A33'] = 'Household Debt Service'
    ws2['A35'] = 'Corporate Profits'
    ws2['A36'] = 'Corporate Debt/GDP'
else:
    ws2 = wb['Sheet2']

def get_fred_latest(series_id):
    """Get latest value from FRED"""
    try:
        data = fred.get_series(series_id)
        data = data.dropna()
        if len(data) > 0:
            return data.iloc[-1]
        return None
    except Exception as e:
        print(f"Error fetching {series_id}: {e}")
        return None

def get_fred_series_monthly(series_id, months=24):
    """Get monthly time series data"""
    try:
        start = datetime.now() - timedelta(days=months*31)
        data = fred.get_series(series_id, observation_start=start)
        data = data.dropna()
        # Resample to monthly (last value of each month)
        if len(data) > 0:
            monthly = data.resample('ME').last().dropna()
            return monthly
        return pd.Series()
    except Exception as e:
        print(f"Error fetching {series_id}: {e}")
        return pd.Series()

print("=" * 50)
print("POPULATING SHEET 1 - Summary Data")
print("=" * 50)

# Fetch all data first
print("Fetching data from FRED...")

gdp = get_fred_latest('GDP')
debt = get_fred_latest('GFDEBTN')
debt_gdp = get_fred_latest('GFDEGDQ188S')
tax_rev = get_fred_latest('FGRECPT')
deficit = get_fred_latest('FYFSD')
ten_yr = get_fred_latest('DGS10')
interest_pmt = get_fred_latest('A091RC1Q027SBEA')
niip = get_fred_latest('IIPUSNETIQ')

# Write to Sheet1 - Using correct row numbers based on Excel structure
print("\nWriting to Sheet1...")

# Row 4: U.S. GDP (B4) - GDP is in billions
if gdp:
    ws1['B4'] = f"${gdp/1000:,.0f} trillion"
    print(f"  GDP: ${gdp/1000:,.2f} trillion")

# Row 5: Total U.S. debt (B5)
if debt:
    ws1['B5'] = f"${debt/1e6:,.0f} trillion"
    print(f"  Debt: ${debt/1e6:,.2f} trillion")

# Row 6: Debt/GDP (B6)
if debt_gdp:
    ws1['B6'] = round(debt_gdp/100, 2)
    print(f"  Debt/GDP: {debt_gdp:.1f}%")

# Row 8: Tax Revenue (B8)
if tax_rev:
    ws1['B8'] = f"${tax_rev:,.0f} billion"
    print(f"  Tax Revenue: ${tax_rev:,.0f} billion")

# Row 9: U.S. Deficit (B9) - deficit is in millions
if deficit:
    ws1['B9'] = f"${abs(deficit)/1000:,.0f} billion"
    print(f"  Deficit: ${abs(deficit)/1000:,.0f} billion")

# Row 10: Borrowed (B10)
if deficit:
    ws1['B10'] = f"${abs(deficit)/1000:,.0f} billion"

# Row 21: Yield on 10 year (B21) - merged cell, use unmerge first
if ten_yr:
    ws1['B21'] = ten_yr/100
    print(f"  10Y Yield: {ten_yr:.2f}%")

# Row 22: Cost to service debt (B22)
if interest_pmt:
    ws1['B22'] = f"~${interest_pmt:,.0f} billion"
    print(f"  Interest Payments: ${interest_pmt:,.0f} billion")

# NIIP data (FRED returns in millions)
if niip:
    niip_trillion = abs(niip) / 1000000
    ws1['D29'] = f"-${niip_trillion:.1f} trillion"
    print(f"  NIIP: -${niip_trillion:.1f} trillion")

print("\n" + "=" * 50)
print("POPULATING SHEET 2 - Time Series Data")
print("=" * 50)

# FRED series mapping for Sheet2
# Format: (row_number, series_id, description)
sheet2_mapping = [
    (3, 'USGOVT', 'Government payrolls'),
    (4, 'USPRIV', 'Private payrolls'),
    (5, 'UNRATE', 'Unemployment rate'),
    (6, 'JTSJOR', 'Job openings rate'),
    (7, 'JTSJOL', 'Job openings level'),
    (8, 'AWHAETP', 'Avg hours worked'),
    (13, 'IC4WSA', 'Initial claims 4W MA'),
    (14, 'CCSA', 'Continuing claims'),
    (15, 'TEMPHELPS', 'Temp jobs'),
    (17, 'RSAFS', 'Retail Sales'),
    (18, 'TOTALSA', 'Auto sales'),
    (19, 'DSPIC96', 'Real disposable income'),
    (20, 'DRCCLACBS', 'Credit card delinquency'),
    (21, 'UMCSENT', 'Consumer sentiment'),
    (22, 'ACTLISCOUUS', 'Housing inventory'),
    (23, 'CSUSHPINSA', 'Case-Shiller'),
    (24, 'MSPUS', 'Median home price'),
    (26, 'MANEMP', 'ISM Mfg Employment'),
    (27, 'NEWORDER', 'Mfg New Orders'),
    (28, 'IPMAN', 'Industrial Production Mfg'),
    (32, 'BOGZ1FL193020005Q', 'Household deposits'),
    (33, 'TDSP', 'Household debt service'),
    (35, 'CP', 'Corporate profits'),
    (36, 'NCBCMDPMVCE', 'Corporate debt/GDP'),
]

# Generate date headers for Sheet2 (24 months back from now)
print("\nWriting date headers...")
current_date = datetime.now()
for col in range(1, 26):
    month_date = current_date - timedelta(days=(col-1)*30)
    ws2.cell(row=2, column=col+1, value=month_date.strftime('%Y-%m-%d'))

# Populate each series
print("\nFetching and writing time series data...")
for row, series_id, desc in sheet2_mapping:
    try:
        data = get_fred_series_monthly(series_id, months=24)
        if len(data) > 0:
            # Write last 24 values (most recent first)
            values = data.tail(24).values[::-1]  # Reverse so most recent is first
            for col, val in enumerate(values[:24], start=2):
                if pd.notna(val):
                    cell = ws2.cell(row=row, column=col, value=round(val, 2))
                    cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            print(f"  Row {row}: {desc} ({series_id}) - {len(values)} values")
        else:
            print(f"  Row {row}: {desc} ({series_id}) - NO DATA")
    except Exception as e:
        print(f"  Row {row}: {desc} ({series_id}) - ERROR: {e}")

# Save workbook
print("\nSaving workbook...")
wb.save(r'C:\Users\daqui\OneDrive\Documents\Targeted Equity Consulting Group\Macro\Macro Data.xlsx')

print("\n" + "=" * 50)
print("DONE! Excel file populated with FRED data.")
print("=" * 50)
