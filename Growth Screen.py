Here's the improved code with 3-year average revenue growth added:

```python
"""
Revenue Analysis Spreadsheet Generator - ENHANCED VERSION
========================================================
- Enhanced: Added 3-year average revenue growth analysis
- Fixed: Revenue ordering (T-4 = oldest to T+2 = newest)
- Fixed: File locking issues (uses timestamp in filename)
- Fixed: API keys from .env file

Requirements:
    pip install yfinance pandas openpyxl requests numpy python-dotenv

Configuration:
    1. Create a .env file with your API keys
    2. Update TICKERS list below
"""

import yfinance as yf
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import requests
from datetime import datetime
import warnings
import os
from dotenv import load_dotenv

warnings.filterwarnings('ignore')

# Load environment variables from .env file
load_dotenv()

# ============================================================================
# CONFIGURATION - EDIT THESE VALUES
# ============================================================================

TICKERS = ['ALAB', 'APH', 'CRDO', 'LITE', 'COHR']  # Add your stock tickers here

# API keys are loaded from .env file
FMP_API_KEY = os.getenv('FMP_API_KEY', '')
FISCAL_API_KEY = os.getenv('FISCAL_API_KEY', '')

# Output file path with timestamp (avoids file locking issues)
timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
OUTPUT_PATH = f'revenue_analysis_{timestamp}.xlsx'


# If you prefer a fixed filename, uncomment this instead:
# OUTPUT_PATH = 'revenue_analysis.xlsx'

# ============================================================================
# DATA FETCHING FUNCTIONS
# ============================================================================

def get_yfinance_data(ticker):
    """
    Fetch financial data from Yahoo Finance (free, no API key required)

    Args:
        ticker (str): Stock ticker symbol

    Returns:
        dict: Contains sector, annual financials, quarterly financials
    """
    try:
        stock = yf.Ticker(ticker)
        info = stock.info
        financials = stock.financials
        quarterly_financials = stock.quarterly_financials
        sector = info.get('sector', '')

        return {
            'sector': sector,
            'annual': financials,
            'quarterly': quarterly_financials,
            'info': info
        }
    except Exception as e:
        print(f"Error fetching yfinance data for {ticker}: {e}")
        return None


def get_fmp_data(ticker, api_key):
    """
    Fetch financial data from Financial Modeling Prep

    Args:
        ticker (str): Stock ticker symbol
        api_key (str): FMP API key

    Returns:
        dict: Contains annual and quarterly income statements
    """
    if not api_key:
        return None

    try:
        # Annual income statement
        annual_url = f'https://financialmodelingprep.com/api/v3/income-statement/{ticker}?period=annual&apikey={api_key}'
        annual_response = requests.get(annual_url, timeout=10)
        annual_data = annual_response.json() if annual_response.status_code == 200 else []

        # Quarterly income statement
        quarterly_url = f'https://financialmodelingprep.com/api/v3/income-statement/{ticker}?period=quarter&apikey={api_key}'
        quarterly_response = requests.get(quarterly_url, timeout=10)
        quarterly_data = quarterly_response.json() if quarterly_response.status_code == 200 else []

        return {
            'annual': annual_data,
            'quarterly': quarterly_data
        }
    except Exception as e:
        print(f"Error fetching FMP data for {ticker}: {e}")
        return None


def get_fiscal_data(ticker, api_key):
    """
    Fetch financial data from Fiscal.ai

    Args:
        ticker (str): Stock ticker symbol
        api_key (str): Fiscal.ai API key

    Returns:
        dict: Contains annual and quarterly financial data
    """
    if not api_key:
        return None

    try:
        headers = {'Authorization': f'Bearer {api_key}'}

        # Annual data
        annual_url = f'https://api.fiscal.ai/v1/financials/{ticker}?period=annual'
        annual_response = requests.get(annual_url, headers=headers, timeout=10)
        annual_data = annual_response.json() if annual_response.status_code == 200 else {}

        # Quarterly data
        quarterly_url = f'https://api.fiscal.ai/v1/financials/{ticker}?period=quarterly'
        quarterly_response = requests.get(quarterly_url, headers=headers, timeout=10)
        quarterly_data = quarterly_response.json() if quarterly_response.status_code == 200 else {}

        return {
            'annual': annual_data,
            'quarterly': quarterly_data
        }
    except Exception as e:
        print(f"Error fetching Fiscal.ai data for {ticker}: {e}")
        return None


def extract_revenue_data(ticker, yf_data, fmp_data, fiscal_data):
    """
    Extract and consolidate revenue data from all sources
    Priority: FMP > Yahoo Finance > Fiscal.ai

    Args:
        ticker (str): Stock ticker symbol
        yf_data (dict): Yahoo Finance data
        fmp_data (dict): Financial Modeling Prep data
        fiscal_data (dict): Fiscal.ai data

    Returns:
        dict: Consolidated revenue data with sector, annual and quarterly revenue
    """
    result = {
        'ticker': ticker,
        'sector': '',
        'annual_revenue': {},
        'quarterly_revenue': {}
    }

    # Extract sector from Yahoo Finance
    if yf_data:
        result['sector'] = yf_data.get('sector', '')

    # Extract annual revenue (prioritize FMP, then yfinance, then Fiscal.ai)
    if fmp_data and fmp_data['annual']:
        for item in fmp_data['annual'][:7]:  # Get last 7 years
            year = item.get('calendarYear', '')
            revenue = item.get('revenue', 0)
            if year and revenue:
                result['annual_revenue'][str(year)] = revenue / 1_000_000  # Convert to millions

    elif yf_data and yf_data['annual'] is not None and not yf_data['annual'].empty:
        try:
            revenue_row = yf_data['annual'].loc['Total Revenue'] if 'Total Revenue' in yf_data['annual'].index else None
            if revenue_row is not None:
                for col in revenue_row.index[:7]:
                    year = col.year if hasattr(col, 'year') else str(col)
                    result['annual_revenue'][str(year)] = revenue_row[col] / 1_000_000
        except Exception as e:
            print(f"Warning: Could not extract annual data for {ticker}: {e}")

    # Extract quarterly revenue (prioritize FMP, then yfinance)
    if fmp_data and fmp_data['quarterly']:
        # FMP returns quarters in reverse chronological order (newest first)
        # We need to map them to Q-4 (oldest) through Q+6 (newest/future)
        for i, item in enumerate(fmp_data['quarterly'][:11]):  # Get last 11 quarters
            period = item.get('period', '')
            revenue = item.get('revenue', 0)
            if revenue:
                # Map: index 0 = most recent (Q+6), index 10 = oldest (Q-4)
                q_index = 6 - i  # Q+6, Q+5, Q+4, ... Q-4
                result['quarterly_revenue'][f"Q{q_index}"] = revenue / 1_000_000

    elif yf_data and yf_data['quarterly'] is not None and not yf_data['quarterly'].empty:
        try:
            revenue_row = yf_data['quarterly'].loc['Total Revenue'] if 'Total Revenue' in yf_data[
                'quarterly'].index else None
            if revenue_row is not None:
                # Yahoo Finance returns in reverse chronological order
                for i, col in enumerate(revenue_row.index[:11]):
                    q_index = 6 - i  # Q+6 down to Q-4
                    result['quarterly_revenue'][f"Q{q_index}"] = revenue_row[col] / 1_000_000
        except Exception as e:
            print(f"Warning: Could not extract quarterly data for {ticker}: {e}")

    return result


# ============================================================================
# SPREADSHEET CREATION FUNCTION
# ============================================================================

def create_spreadsheet(data_dict):
    """
    Create comprehensive Excel spreadsheet with revenue analysis

    Structure:
        - Left side: Annual revenue and growth with rankings
        - Right side: Quarterly revenue and growth with rankings
        - Added: 3-Year Average Revenue Growth analysis

    Args:
        data_dict (dict): Dictionary of ticker data

    Returns:
        Workbook: Excel workbook object
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Revenue Analysis"

    # Define styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    subheader_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')

    current_row = 1

    # ========================================================================
    # SECTION 1: Annual Revenue
    # ========================================================================
    ws.merge_cells(f'A{current_row}:I{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = 'Annual revenue'
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    current_row += 1

    # Headers
    headers = ['Ticker', 'sector', 'revenue T-4', 'revenue T-3', 'revenue T-2',
               'revenue T-1', 'revenue T-0', 'revenue T+1', 'revenue T+2']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.border = border

    start_annual_data_row = current_row + 1

    # Data rows
    for ticker in TICKERS:
        current_row += 1
        ws.cell(row=current_row, column=1, value=ticker)

        if ticker in data_dict:
            ws.cell(row=current_row, column=2, value=data_dict[ticker]['sector'])

            # Get annual revenue sorted by year (oldest to newest: T-4 to T+2)
            annual_rev = data_dict[ticker]['annual_revenue']
            years = sorted(annual_rev.keys())  # Ascending order: oldest first

            for i, year in enumerate(years[:7]):
                col = 3 + i
                if col <= 9:
                    ws.cell(row=current_row, column=col, value=annual_rev[year])
                    ws.cell(row=current_row, column=col).number_format = '#,##0'

    end_annual_data_row = current_row
    current_row += 2

    # ========================================================================
    # SECTION 2: Annual Revenue Growth
    # ========================================================================
    ws.merge_cells(f'A{current_row}:I{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = 'Annual Revenue Growth'
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    current_row += 1

    # Headers
    growth_headers = ['', 'sector', 'revenue T-4', 'revenue T-3', 'revenue T-2',
                      'revenue T-1', 'revenue T-0', 'revenue T+1', 'revenue T+2']
    for col, header in enumerate(growth_headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.border = border

    start_growth_row = current_row + 1

    # Growth calculations using Excel formulas
    for i, ticker in enumerate(TICKERS):
        current_row += 1
        ws.cell(row=current_row, column=1, value=ticker)

        if ticker in data_dict:
            ws.cell(row=current_row, column=2, value=data_dict[ticker]['sector'])

            # Calculate YoY growth rates using formulas
            data_row = start_annual_data_row + i
            for col in range(4, 10):  # columns D through I
                col_letter = get_column_letter(col)
                prev_col_letter = get_column_letter(col - 1)
                formula = f'=IF(AND({prev_col_letter}{data_row}<>0, {col_letter}{data_row}<>0), ({col_letter}{data_row}-{prev_col_letter}{data_row})/{prev_col_letter}{data_row}, "")'
                ws.cell(row=current_row, column=col, value=formula)
                ws.cell(row=current_row, column=col).number_format = '0.0%'

    end_growth_row = current_row
    current_row += 2

    # ========================================================================
    # SECTION 3: 3-Year Average Revenue Growth (NEW FEATURE)
    # ========================================================================
    ws.merge_cells(f'A{current_row}:H{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = '3-Year Average Revenue Growth'
    cell.font = header_font
    cell.fill = subheader_fill
    cell.alignment = center_align
    current_row += 1

    # Headers for 3-year averages
    avg_headers = ['Ticker', 'Sector', 'Avg Growth T-4 to T-1', 'Avg Growth T-3 to T-0', 
                   'Avg Growth T-2 to T+1', 'Avg Growth T-1 to T+2', 
                   'Historical 3Yr Avg (T-4 to T-1)', 'Forward 3Yr Avg (T-0 to T+2)']
    for col, header in enumerate(avg_headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.border = border

    start_3yr_avg_row = current_row + 1

    # Calculate 3-year averages using Excel formulas
    for i, ticker in enumerate(TICKERS):
        current_row += 1
        ws.cell(row=current_row, column=1, value=ticker)

        if ticker in data_dict:
            ws.cell(row=current_row, column=2, value=data_dict[ticker]['sector'])

            growth_row = start_growth_row + i

            # 3-year rolling averages
            # T-4 to T-1 average (columns D to F in growth section)
            formula1 = f'=AVERAGE(D{growth_row}:F{growth_row})'
            ws.cell(row=current_row, column=3, value=formula1)
            ws.cell(row=current_row, column=3).number_format = '0.0%'

            # T-3 to T-0 average (columns E to G in growth section)
            formula2