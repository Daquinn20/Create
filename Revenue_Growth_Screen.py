import yfinance as yf
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import time
import os

# Check if dotenv is available
try:
    from dotenv import load_dotenv

    # Load environment variables from .env file
    load_dotenv()
    print("✓ dotenv loaded successfully")
except ImportError:
    print("⚠️  python-dotenv not installed. Please run: pip install python-dotenv")
    print("    Continuing without .env file support...")


    def load_dotenv():
        pass

# API Keys - Automatically loaded from .env file
FMP_API_KEY = os.getenv('FMP_API_KEY')
FISCAL_AI_KEY = os.getenv('FISCAL_AI_KEY')

print(f"Starting Revenue Growth Screen...")


def calculate_cagr(start_value, end_value, periods):
    """Calculate Compound Annual Growth Rate"""
    try:
        if start_value <= 0 or end_value <= 0:
            return None
        cagr = (pow(end_value / start_value, 1 / periods) - 1) * 100
        return round(cagr, 2)
    except:
        return None


def calculate_sequential_growth(current, previous):
    """Calculate sequential quarter-over-quarter growth %"""
    try:
        if not current or not previous or current == 'N/A' or previous == 'N/A':
            return 'N/A'
        if previous == 0:
            return 'N/A'
        growth = ((current - previous) / previous) * 100
        return round(growth, 2)
    except:
        return 'N/A'


def get_fmp_revenue_data(ticker):
    """
    Fetch comprehensive revenue data from FMP:
    - Last 5 years annual revenue
    - Last 6 quarters revenue
    - Forward revenue estimates
    - Calculate sequential growth
    """
    try:
        base_url = "https://financialmodelingprep.com/api/v3"

        # Get company profile for basic info
        profile_url = f"{base_url}/profile/{ticker}?apikey={FMP_API_KEY}"
        profile_response = requests.get(profile_url, timeout=10)

        if profile_response.status_code != 200:
            print(f"  FMP API error for {ticker}: Status {profile_response.status_code}")
            return None

        profile_data = profile_response.json()
        if not profile_data or len(profile_data) == 0:
            print(f"  No FMP profile data for {ticker}")
            return None

        profile = profile_data[0]

        # Get annual income statements (last 5 years)
        annual_url = f"{base_url}/income-statement/{ticker}?limit=5&apikey={FMP_API_KEY}"
        annual_response = requests.get(annual_url, timeout=10)
        annual_data = annual_response.json() if annual_response.status_code == 200 else []

        # Get quarterly income statements (last 8 quarters to be safe)
        quarterly_url = f"{base_url}/income-statement/{ticker}?period=quarter&limit=8&apikey={FMP_API_KEY}"
        quarterly_response = requests.get(quarterly_url, timeout=10)
        quarterly_data = quarterly_response.json() if quarterly_response.status_code == 200 else []

        # Get analyst estimates
        estimates_url = f"{base_url}/analyst-estimates/{ticker}?limit=8&apikey={FMP_API_KEY}"
        estimates_response = requests.get(estimates_url, timeout=10)
        estimates_data = estimates_response.json() if estimates_response.status_code == 200 else []

        # Get revenue estimates by period
        revenue_estimates_url = f"{base_url}/analyst-estimates/{ticker}?period=quarter&limit=10&apikey={FMP_API_KEY}"
        rev_est_response = requests.get(revenue_estimates_url, timeout=10)
        revenue_estimates = rev_est_response.json() if rev_est_response.status_code == 200 else []

        data = {
            'ticker': ticker,
            'company_name': profile.get('companyName', 'N/A'),
            'source': 'FMP'
        }

        # Process Annual Revenue (Last 5 Years)
        if annual_data and len(annual_data) >= 2:
            annual_data_sorted = sorted(annual_data, key=lambda x: x.get('date', ''), reverse=False)

            for i, year_data in enumerate(annual_data_sorted[-5:], 1):
                data[f'revenue_y{i}'] = year_data.get('revenue', 'N/A')
                data[f'revenue_y{i}_date'] = year_data.get('date', 'N/A')

            # Calculate CAGR
            if len(annual_data_sorted) >= 2:
                # 1-Year CAGR
                if len(annual_data_sorted) >= 2:
                    recent_rev = annual_data_sorted[-1].get('revenue')
                    one_year_ago_rev = annual_data_sorted[-2].get('revenue')
                    data['cagr_1y'] = calculate_cagr(one_year_ago_rev, recent_rev, 1)

                # 3-Year CAGR
                if len(annual_data_sorted) >= 4:
                    recent_rev = annual_data_sorted[-1].get('revenue')
                    three_years_ago_rev = annual_data_sorted[-4].get('revenue')
                    data['cagr_3y'] = calculate_cagr(three_years_ago_rev, recent_rev, 3)

                # 5-Year CAGR
                if len(annual_data_sorted) >= 5:
                    recent_rev = annual_data_sorted[-1].get('revenue')
                    five_years_ago_rev = annual_data_sorted[0].get('revenue')
                    data['cagr_5y'] = calculate_cagr(five_years_ago_rev, recent_rev, 5)

        # Process Quarterly Revenue (Last 6 Quarters)
        if quarterly_data and len(quarterly_data) >= 1:
            quarterly_sorted = sorted(quarterly_data, key=lambda x: x.get('date', ''), reverse=False)

            # Store revenues for sequential growth calculation
            revenues = []
            for i, qtr_data in enumerate(quarterly_sorted[-6:], 1):
                rev = qtr_data.get('revenue', 'N/A')
                data[f'revenue_q{i}'] = rev
                data[f'revenue_q{i}_date'] = qtr_data.get('date', 'N/A')
                revenues.append(rev)

            # Calculate sequential growth (Q-over-Q)
            # Q-5 vs Q-6, Q-4 vs Q-5, Q-3 vs Q-4, Q-2 vs Q-3, Q-1 vs Q-2
            for i in range(1, 6):
                if i < len(revenues):
                    current = revenues[i]
                    previous = revenues[i - 1]
                    data[f'q{i}_sequential_growth'] = calculate_sequential_growth(current, previous)

        # Process Forward Estimates
        if estimates_data:
            estimates_sorted = sorted(estimates_data, key=lambda x: x.get('date', ''), reverse=False)

            # Forward annual estimates
            last_annual_date = annual_data_sorted[-1].get('date', '') if annual_data_sorted else ''
            future_estimates = [e for e in estimates_sorted if e.get('date', '') > last_annual_date]

            if len(future_estimates) >= 1:
                data['est_revenue_fy1'] = future_estimates[0].get('estimatedRevenueAvg', 'N/A')
                data['est_revenue_fy1_date'] = future_estimates[0].get('date', 'N/A')

            if len(future_estimates) >= 2:
                data['est_revenue_fy2'] = future_estimates[1].get('estimatedRevenueAvg', 'N/A')
                data['est_revenue_fy2_date'] = future_estimates[1].get('date', 'N/A')

                # Calculate forward 2-year growth
                fy1_rev = future_estimates[0].get('estimatedRevenueAvg')
                fy2_rev = future_estimates[1].get('estimatedRevenueAvg')
                if fy1_rev and fy2_rev and fy1_rev > 0:
                    fwd_2y_growth = ((fy2_rev - fy1_rev) / fy1_rev) * 100
                    data['forward_2y_growth'] = round(fwd_2y_growth, 2)

        # Process Forward Quarterly Estimates (Next 4 Quarters)
        if revenue_estimates:
            rev_est_sorted = sorted(revenue_estimates, key=lambda x: x.get('date', ''), reverse=False)

            # Get future quarters only
            last_actual_date = quarterly_sorted[-1].get('date', '') if quarterly_sorted else ''
            future_qtrs = [e for e in rev_est_sorted if e.get('date', '') > last_actual_date]

            # Store forward revenues
            fwd_revenues = []
            for i, qtr_est in enumerate(future_qtrs[:4], 1):
                est_rev = qtr_est.get('estimatedRevenueAvg', 'N/A')
                data[f'est_revenue_fq{i}'] = est_rev
                data[f'est_revenue_fq{i}_date'] = qtr_est.get('date', 'N/A')
                fwd_revenues.append(est_rev)

            # Calculate forward sequential growth
            # Q+1 vs Q-1, Q+2 vs Q+1, Q+3 vs Q+2, Q+4 vs Q+3
            if revenues and len(revenues) > 0:
                # Q+1 growth vs Q-1 (most recent actual)
                last_actual = revenues[-1] if len(revenues) > 0 else None
                if last_actual and len(fwd_revenues) > 0:
                    data['fq1_sequential_growth'] = calculate_sequential_growth(fwd_revenues[0], last_actual)

            # Q+2 vs Q+1, Q+3 vs Q+2, Q+4 vs Q+3
            for i in range(1, 4):
                if i < len(fwd_revenues):
                    data[f'fq{i + 1}_sequential_growth'] = calculate_sequential_growth(fwd_revenues[i],
                                                                                       fwd_revenues[i - 1])

        print(f"  ✓ FMP revenue data retrieved for {ticker}")
        return data

    except Exception as e:
        print(f"  FMP error for {ticker}: {str(e)}")
        return None


def get_fiscalai_revenue_data(ticker):
    """
    Fetch financial data from Fiscal.ai (secondary source after FMP).
    Fiscal.ai provides institutional-grade financial data with deep coverage.
    """
    try:
        if not FISCAL_AI_KEY:
            return None

        base_url = "https://api.fiscal.ai/v1"
        headers = {
            "X-API-KEY": FISCAL_AI_KEY,
            "Content-Type": "application/json"
        }

        # Get company profile/info
        profile_url = f"{base_url}/company/{ticker}/profile"
        profile_response = requests.get(profile_url, headers=headers, timeout=10)

        if profile_response.status_code != 200:
            print(f"  Fiscal.ai API error for {ticker}: Status {profile_response.status_code}")
            return None

        profile_data = profile_response.json()

        # Get financial statements (annual and quarterly)
        financials_url = f"{base_url}/financials/{ticker}"
        financials_response = requests.get(financials_url, headers=headers, timeout=10)

        if financials_response.status_code != 200:
            print(f"  Fiscal.ai financials error for {ticker}")
            return None

        financials_data = financials_response.json()

        # Get analyst estimates
        estimates_url = f"{base_url}/estimates/{ticker}"
        estimates_response = requests.get(estimates_url, headers=headers, timeout=10)
        estimates_data = estimates_response.json() if estimates_response.status_code == 200 else {}

        data = {
            'ticker': ticker,
            'company_name': profile_data.get('name', 'N/A'),
            'source': 'Fiscal.ai'
        }

        # Process annual and quarterly revenue data
        annual_statements = financials_data.get('annual', [])
        quarterly_statements = financials_data.get('quarterly', [])

        # Process last 5 years annual revenue
        for i, stmt in enumerate(annual_statements[-5:], 1):
            data[f'revenue_y{i}'] = stmt.get('revenue', 'N/A')
            data[f'revenue_y{i}_date'] = stmt.get('report_period', 'N/A')

        # Calculate CAGR if enough data
        if len(annual_statements) >= 2:
            recent_rev = annual_statements[-1].get('revenue')
            one_year_ago = annual_statements[-2].get('revenue')
            data['cagr_1y'] = calculate_cagr(one_year_ago, recent_rev, 1)

            if len(annual_statements) >= 4:
                three_years_ago = annual_statements[-4].get('revenue')
                data['cagr_3y'] = calculate_cagr(three_years_ago, recent_rev, 3)

            if len(annual_statements) >= 5:
                five_years_ago = annual_statements[0].get('revenue')
                data['cagr_5y'] = calculate_cagr(five_years_ago, recent_rev, 5)

        # Process last 6 quarters with sequential growth
        revenues = []
        for i, stmt in enumerate(quarterly_statements[-6:], 1):
            rev = stmt.get('revenue', 'N/A')
            data[f'revenue_q{i}'] = rev
            data[f'revenue_q{i}_date'] = stmt.get('report_period', 'N/A')
            revenues.append(rev)

        # Calculate sequential growth
        for i in range(1, 6):
            if i < len(revenues):
                data[f'q{i}_sequential_growth'] = calculate_sequential_growth(revenues[i], revenues[i - 1])

        # Process forward estimates
        if estimates_data:
            annual_estimates = estimates_data.get('annual', [])
            quarterly_estimates = estimates_data.get('quarterly', [])

            if len(annual_estimates) >= 1:
                data['est_revenue_fy1'] = annual_estimates[0].get('revenue_estimate', 'N/A')
                data['est_revenue_fy1_date'] = annual_estimates[0].get('period', 'N/A')

            if len(annual_estimates) >= 2:
                data['est_revenue_fy2'] = annual_estimates[1].get('revenue_estimate', 'N/A')
                data['est_revenue_fy2_date'] = annual_estimates[1].get('period', 'N/A')

                # Calculate forward 2-year growth
                fy1 = annual_estimates[0].get('revenue_estimate')
                fy2 = annual_estimates[1].get('revenue_estimate')
                if fy1 and fy2 and fy1 > 0:
                    fwd_2y_growth = ((fy2 - fy1) / fy1) * 100
                    data['forward_2y_growth'] = round(fwd_2y_growth, 2)

            # Process forward quarterly estimates with sequential growth
            fwd_revenues = []
            for i, qtr in enumerate(quarterly_estimates[:4], 1):
                est_rev = qtr.get('revenue_estimate', 'N/A')
                data[f'est_revenue_fq{i}'] = est_rev
                data[f'est_revenue_fq{i}_date'] = qtr.get('period', 'N/A')
                fwd_revenues.append(est_rev)

            # Calculate forward sequential growth
            if revenues and len(revenues) > 0:
                last_actual = revenues[-1]
                if len(fwd_revenues) > 0:
                    data['fq1_sequential_growth'] = calculate_sequential_growth(fwd_revenues[0], last_actual)

            for i in range(1, 4):
                if i < len(fwd_revenues):
                    data[f'fq{i + 1}_sequential_growth'] = calculate_sequential_growth(fwd_revenues[i],
                                                                                       fwd_revenues[i - 1])

        print(f"  ✓ Fiscal.ai revenue data retrieved for {ticker}")
        return data

    except Exception as e:
        print(f"  Fiscal.ai error for {ticker}: {str(e)}")
        return None


def get_yfinance_revenue_data(ticker):
    """
    Fallback: Fetch revenue data using yfinance
    Note: yfinance has limited forward estimates
    """
    try:
        stock = yf.Ticker(ticker)
        info = stock.info

        # Get annual financials
        annual_financials = stock.financials
        quarterly_financials = stock.quarterly_financials

        data = {
            'ticker': ticker,
            'company_name': info.get('longName', 'N/A'),
            'source': 'yfinance'
        }

        # Extract revenue from financials if available
        if annual_financials is not None and not annual_financials.empty:
            if 'Total Revenue' in annual_financials.index:
                revenues = annual_financials.loc['Total Revenue'].dropna()
                for i, (date, revenue) in enumerate(list(revenues.items())[:5], 1):
                    data[f'revenue_y{i}'] = revenue
                    data[f'revenue_y{i}_date'] = str(date.date())

        if quarterly_financials is not None and not quarterly_financials.empty:
            if 'Total Revenue' in quarterly_financials.index:
                revenues_list = []
                rev_data = quarterly_financials.loc['Total Revenue'].dropna()
                for i, (date, revenue) in enumerate(list(rev_data.items())[:6], 1):
                    data[f'revenue_q{i}'] = revenue
                    data[f'revenue_q{i}_date'] = str(date.date())
                    revenues_list.append(revenue)

                # Calculate sequential growth
                for i in range(1, min(6, len(revenues_list))):
                    data[f'q{i}_sequential_growth'] = calculate_sequential_growth(revenues_list[i],
                                                                                  revenues_list[i - 1])

        # yfinance has limited forward data
        data['est_revenue_fy1'] = info.get('revenueEstimate', {}).get('avg', 'N/A') if isinstance(
            info.get('revenueEstimate'), dict) else 'N/A'

        print(f"  ✓ yfinance revenue data retrieved for {ticker}")
        return data

    except Exception as e:
        print(f"  yfinance error for {ticker}: {str(e)}")
        return None


def get_stock_revenue_data(ticker):
    """
    Get revenue data with prioritized fallback:
    1. Try FMP (most trusted - comprehensive data)
    2. Try Fiscal.ai (institutional-grade data)
    3. Fall back to yfinance (free but limited)
    """
    print(f"Processing {ticker}...")

    # Try FMP first (most trusted)
    data = get_fmp_revenue_data(ticker)
    if data:
        return data

    print(f"  Falling back to Fiscal.ai for {ticker}...")
    # Try Fiscal.ai second
    data = get_fiscalai_revenue_data(ticker)
    if data:
        return data

    print(f"  Falling back to yfinance for {ticker}...")
    # Fall back to yfinance
    data = get_yfinance_revenue_data(ticker)
    if data:
        return data

    print(f"✗ All data sources failed for {ticker}")
    return None


def create_revenue_spreadsheet(data_dict):
    """Create an Excel spreadsheet with four-section layout including sequential growth"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Revenue Growth Screen"

    # SECTION 1: Annual Revenue History + Forward Estimates + Data Source
    annual_headers = [
        'Ticker', 'Company Name',
        'Y-5 Revenue', 'Y-4 Revenue', 'Y-3 Revenue', 'Y-2 Revenue', 'Y-1 Revenue',
        'FY+1 Est Rev', 'FY+2 Est Rev', 'Data Source'
    ]

    # SECTION 2: CAGR Metrics
    cagr_headers = [
        'Ticker', 'Company Name',
        '5Y CAGR %', '3Y CAGR %', '1Y CAGR %', 'Fwd 2Y Growth %'
    ]

    # SECTION 3: Quarterly Revenue Data
    quarterly_headers = [
        'Ticker', 'Company Name',
        'Q-6 Revenue', 'Q-5 Revenue', 'Q-4 Revenue', 'Q-3 Revenue', 'Q-2 Revenue', 'Q-1 Revenue',
        'FQ+1 Est Rev', 'FQ+2 Est Rev', 'FQ+3 Est Rev', 'FQ+4 Est Rev'
    ]

    # SECTION 4: Sequential Growth (Quarter-over-Quarter)
    sequential_headers = [
        'Ticker', 'Company Name',
        'Q-5 Rev Sequential Growth', 'Q-4 Rev Sequential Growth', 'Q-3 Rev Sequential Growth',
        'Q-2 Rev Sequential Growth', 'Q-1 Rev Sequential Growth', 'Q+1 Rev Sequential Growth',
        'Q+2 Rev Sequential Growth', 'Q+3 Rev Sequential Growth', 'Q+4 Rev Sequential Growth'
    ]

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    current_row = 1

    # ============= COMPANY HEADER ROW =============
    # Merge cells for company name header
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(annual_headers))
    header_cell = ws.cell(row=current_row, column=1)
    header_cell.value = "Targeted Equity Consulting Group"
    header_cell.font = Font(bold=True, size=14, color="FFFFFF")
    header_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_cell.alignment = Alignment(horizontal="center", vertical="center")

    current_row += 1

    # ============= SECTION 1: ANNUAL REVENUE =============
    for col_num, header in enumerate(annual_headers, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    current_row += 1
    for ticker, data in data_dict.items():
        col = 1
        ws.cell(row=current_row, column=col, value=data.get('ticker', 'N/A'));
        col += 1
        ws.cell(row=current_row, column=col, value=data.get('company_name', 'N/A'));
        col += 1

        # Y-5 through Y-1
        for i in range(1, 6):
            ws.cell(row=current_row, column=col, value=data.get(f'revenue_y{i}', 'N/A'));
            col += 1

        # Forward estimates
        ws.cell(row=current_row, column=col, value=data.get('est_revenue_fy1', 'N/A'));
        col += 1
        ws.cell(row=current_row, column=col, value=data.get('est_revenue_fy2', 'N/A'));
        col += 1
        ws.cell(row=current_row, column=col, value=data.get('source', 'N/A'));
        col += 1

        for col_num in range(1, len(annual_headers) + 1):
            ws.cell(row=current_row, column=col_num).border = thin_border
            ws.cell(row=current_row, column=col_num).alignment = Alignment(horizontal="center", vertical="center")

        current_row += 1

    current_row += 1

    # ============= SECTION 2: CAGR METRICS =============
    for col_num, header in enumerate(cagr_headers, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    current_row += 1
    for ticker, data in data_dict.items():
        col = 1
        ws.cell(row=current_row, column=col, value=data.get('ticker', 'N/A'));
        col += 1
        ws.cell(row=current_row, column=col, value=data.get('company_name', 'N/A'));
        col += 1
        ws.cell(row=current_row, column=col, value=data.get('cagr_5y', 'N/A'));
        col += 1
        ws.cell(row=current_row, column=col, value=data.get('cagr_3y', 'N/A'));
        col += 1
        ws.cell(row=current_row, column=col, value=data.get('cagr_1y', 'N/A'));
        col += 1
        ws.cell(row=current_row, column=col, value=data.get('forward_2y_growth', 'N/A'));
        col += 1

        for col_num in range(1, len(cagr_headers) + 1):
            ws.cell(row=current_row, column=col_num).border = thin_border
            ws.cell(row=current_row, column=col_num).alignment = Alignment(horizontal="center", vertical="center")

        current_row += 1

    current_row += 1

    # ============= SECTION 3: QUARTERLY DATA =============
    for col_num, header in enumerate(quarterly_headers, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    current_row += 1
    for ticker, data in data_dict.items():
        col = 1
        ws.cell(row=current_row, column=col, value=data.get('ticker', 'N/A'));
        col += 1
        ws.cell(row=current_row, column=col, value=data.get('company_name', 'N/A'));
        col += 1

        # Q-6 through Q-1
        for i in range(1, 7):
            ws.cell(row=current_row, column=col, value=data.get(f'revenue_q{i}', 'N/A'));
            col += 1

        # Forward quarterly estimates
        for i in range(1, 5):
            ws.cell(row=current_row, column=col, value=data.get(f'est_revenue_fq{i}', 'N/A'));
            col += 1

        for col_num in range(1, len(quarterly_headers) + 1):
            ws.cell(row=current_row, column=col_num).border = thin_border
            ws.cell(row=current_row, column=col_num).alignment = Alignment(horizontal="center", vertical="center")

        current_row += 1

    current_row += 1

    # ============= SECTION 4: SEQUENTIAL GROWTH =============
    for col_num, header in enumerate(sequential_headers, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    current_row += 1
    for ticker, data in data_dict.items():
        col = 1
        ws.cell(row=current_row, column=col, value=data.get('ticker', 'N/A'));
        col += 1
        ws.cell(row=current_row, column=col, value=data.get('company_name', 'N/A'));
        col += 1

        # Sequential growth Q-5 through Q-1
        for i in range(1, 6):
            ws.cell(row=current_row, column=col, value=data.get(f'q{i}_sequential_growth', 'N/A'));
            col += 1

        # Forward sequential growth Q+1 through Q+4
        for i in range(1, 5):
            ws.cell(row=current_row, column=col, value=data.get(f'fq{i}_sequential_growth', 'N/A'));
            col += 1

        for col_num in range(1, len(sequential_headers) + 1):
            ws.cell(row=current_row, column=col_num).border = thin_border
            ws.cell(row=current_row, column=col_num).alignment = Alignment(horizontal="center", vertical="center")

        current_row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30

    for col in range(3,
                     max(len(annual_headers), len(cagr_headers), len(quarterly_headers), len(sequential_headers)) + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 15

    return wb


def main():
    print("=" * 70)
    print("REVENUE GROWTH SCREEN - Complete Analysis")
    print("Four Sections: Annual | CAGR | Quarterly | Sequential Growth")
    print("Data Priority: FMP > Fiscal.ai > yfinance")
    print("=" * 70)

    # Check if API keys are loaded
    if not FMP_API_KEY:
        print("\n⚠️  WARNING: FMP_API_KEY not found in .env file!")
        print("   Please create a .env file in the same directory with:")
        print("   FMP_API_KEY=your_key_here")
        print("   Get your free key at: https://financialmodelingprep.com/developer/docs/")
    else:
        print(f"\n✓ FMP API key loaded successfully")

    if FISCAL_AI_KEY:
        print(f"✓ Fiscal.ai API key loaded successfully")
    else:
        print("⚠️  Fiscal.ai key not found - will skip this source")

    print("✓ yfinance available as final fallback")
    print()

    # List of tickers to analyze
    tickers = ['ALAB', 'CRDO', 'COHR', 'LITE', 'AAOI']

    print(f"\nAnalyzing {len(tickers)} tickers: {tickers}\n")

    data_dict = {}
    successful = 0
    sources_used = {'FMP': 0, 'Fiscal.ai': 0, 'yfinance': 0}

    for ticker in tickers:
        data = get_stock_revenue_data(ticker)
        if data:
            data_dict[ticker] = data
            successful += 1
            source = data.get('source', 'Unknown')
            sources_used[source] = sources_used.get(source, 0) + 1

        # Rate limiting
        time.sleep(0.8)

    print("\n" + "=" * 70)
    print(f"Data Collection Summary:")
    print(f"  Successfully collected: {successful}/{len(tickers)} tickers")
    print(f"  Data sources used:")
    for source, count in sources_used.items():
        if count > 0:
            print(f"    - {source}: {count} tickers")
    print("=" * 70)

    if data_dict:
        print("\nCreating Excel spreadsheet...")
        wb = create_revenue_spreadsheet(data_dict)

        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"revenue_growth_screen_{timestamp}.xlsx"

        wb.save(filename)
        print(f"\n✓ Revenue Growth Screen saved as: {filename}")
        print("\nSpreadsheet Layout:")
        print("  📊 Section 1: Annual Revenue (Y-5 to Y-1) + Forward (FY+1, FY+2) + Data Source")
        print("  📈 Section 2: CAGR Metrics (5Y, 3Y, 1Y) + Forward 2Y Growth")
        print("  📅 Section 3: Quarterly Revenue (Q-6 to Q-1) + Forward (FQ+1 to FQ+4)")
        print("  📊 Section 4: Sequential Growth % (Q-over-Q for historical & forward)")
    else:
        print("\n✗ No data collected. Spreadsheet not created.")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        import traceback

        traceback.print_exc()
        input("\nPress Enter to exit...")